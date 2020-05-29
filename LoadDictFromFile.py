import codecs
import csv
import os
import re
import unittest

import openpyxl
import xlrd

import GlobalFunctions
from GlobalFunctions import print  # Todo use logging


def __correct(value, optimize):
    if value is None: return ''
    if optimize and isinstance(value, str): return re.sub(r'\s+', ' ', value).strip()
    if isinstance(value, float) and str(value).endswith('.0'): return str(value)[:-2]
    return str(value)


def __titles(titles_original, language, optimize):
    titles = []
    for each in titles_original:
        new_name1 = each[each.find('<') + 1:each.find('>')] if ('<' in each and '>' in each) else each
        new_name2 = each[each.find('(') + 1:each.find(')')] if ('(' in each and ')' in each) else ''
        if new_name2 == language: new_name1 += '_' + new_name2
        titles.append(__correct(new_name1, optimize))
    return titles


def __find_index(maincolumn, titles):
    if maincolumn is None and 'sku' in titles:
        return titles.index('sku')
    elif maincolumn in titles:
        return titles.index(maincolumn)
    return None


def __xls_titles(sheet, optimize, title_row):
    titles_original = []
    for column in range(0, sheet.ncols):
        if title_row is not None:
            data = str(__correct(sheet.cell(title_row - 1, column).value, optimize))
        else:
            data = openpyxl.utils.get_column_letter(column + 1)
        if data == '': data = openpyxl.utils.get_column_letter(column + 1)
        titles_original.append(data)
    return titles_original


def __xlsx_titles(sheet, optimize, title_row):
    titles_original = []
    for column in range(1, sheet.max_column + 1):
        if title_row is not None:
            data = str(__correct(sheet.cell(row=title_row, column=column).value, optimize))
        else:
            data = openpyxl.utils.get_column_letter(column)
        if data == '': data = openpyxl.utils.get_column_letter(column)
        titles_original.append(data)
    return titles_original


def alphabet(length):
    result = []
    for i in range(length):
        result.append(openpyxl.utils.get_column_letter(i + 1))
    return result


@GlobalFunctions.print_run_time
def _csv_import(filename, maincolumn, language, optimize, recognize, delimiter, title_row, first_row):
    res = {}
    with codecs.open(filename, 'r', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter, quotechar='"')
        data = [row for row in reader]
    titles = __titles(data[title_row - 1], language, optimize) if title_row else alphabet(len(data[0]))
    index = __find_index(maincolumn, titles)
    for a, row in enumerate(data[first_row - 1:]):
        if not len(row): continue
        name = str(__correct(row[index], optimize) if index is not None else a + 2)
        if name: res[name] = {titles[i]: __correct(row[i], optimize) for i in range(0, len(titles))}
    print(f"{filename} / {len(data) - 1} lines / {len(res)} loaded / {len(data) - 1 - len(res)} lost / ", end='')
    if recognize: _recognize_data(res)
    return res


@GlobalFunctions.print_run_time
def _xls_import(filename, maincolumn, language, optimize, recognize, title_row, first_row):
    res = {}
    sheet = xlrd.open_workbook(filename).sheet_by_index(0)  # Todo
    titles = __titles(__xls_titles(sheet, optimize, title_row), language, optimize)
    index = __find_index(maincolumn, titles)
    first_row = max(first_row, title_row + 1) if title_row else first_row
    for a in range(first_row - 1, sheet.nrows):
        row = [__correct(__xlrd_get_value(sheet.cell(a, col)), optimize) for col in range(0, len(titles))]
        name = str(row[index] if index is not None else a + 1)
        if name: res[name] = {titles[i]: row[i] for i in range(0, len(titles))}
    rows_count = sheet.nrows - title_row if title_row else sheet.nrows
    print(f"{filename} / {rows_count} lines / {len(res)} loaded / {rows_count - len(res)} lost / ", end='')
    if recognize: _recognize_data(res)
    return res


def __xlrd_get_value(cell):
    if cell.ctype == xlrd.XL_CELL_ERROR: return ''
    return cell.value


@GlobalFunctions.print_run_time
def _xlsx_import(filename, maincolumn, language, optimize, recognize, title_row, first_row):
    res = {}
    sheet = openpyxl.load_workbook(filename).active
    titles = __titles(__xlsx_titles(sheet, optimize, title_row), language, optimize)
    index = __find_index(maincolumn, titles)
    first_row = max(first_row, title_row + 1) if title_row else first_row
    for a in range(first_row, sheet.max_row + 1):
        row = [__correct(sheet.cell(row=a, column=col).value, optimize) for col in range(1, len(titles) + 1)]
        name = str(row[index] if index is not None else a)
        if name: res[name] = {titles[i]: row[i] for i in range(0, len(titles))}
    rows_count = sheet.max_row - title_row if title_row else sheet.max_row
    print(f"{filename} / {rows_count} lines / {len(res)} loaded / {rows_count - len(res)} lost / ", end='')
    if recognize: _recognize_data(res)
    return res


def load(filename, maincolumn=None, language=None, optimize=False, recognize=False, delimiter=',', title_row=1,
         first_row=2):
    if filename.endswith('.csv'):
        return _csv_import(filename, maincolumn, language, optimize, recognize, delimiter, title_row, first_row)
    elif filename.endswith('.xls'):
        return _xls_import(filename, maincolumn, language, optimize, recognize, title_row, first_row)
    elif filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        try:
            return _xlsx_import(filename, maincolumn, language, optimize, recognize, title_row, first_row)
        except KeyError:
            print('Error: bad file format. Will try to use xls instead')
            return _xls_import(filename, maincolumn, language, optimize, recognize, title_row, first_row)
    else:
        raise ValueError(f'Wrong filetype: {filename}')


def change_main_column(data, maincolumn):
    result = {}
    for each in data.values(): result[each[maincolumn]] = each
    print(f'change_main_column: {len(data)} lines / {len(result)} loaded / {len(data) - len(result)} lost')
    return result


def count_not_empty_values(data, column):
    result = 0
    for each in data.values():
        value = each.get(column, None)
        if value is not None and value != '':
            result += 1
    return result


def count_values(data, column, value):
    result = 0
    for each in data.values():
        if each.get(column, '') == value:
            result += 1
    return result


def find_value(data, key, value):
    for each in data.values():
        found_value = each.get(key, '')
        if found_value == value: return each


def _recognize_data(data):
    from CheckTypes import CheckTypesTry
    types, titles = {}, set()
    for row in data.values():
        for key, value in row.items():
            if key not in titles:
                titles.add(key)
                types[key] = int
            if types[key] == str:
                continue
            elif types[key] == int and CheckTypesTry.isint(value):
                continue
            elif types[key] != str and CheckTypesTry.isfloat(value):
                types[key] = float
            else:
                types[key] = str
    for row in data.values():
        for key, value in row.items():
            if types[key] == str:
                continue
            elif types[key] == int and value != '':
                row[key] = int(value)
            elif types[key] == float and value != '':
                row[key] = float(value.replace(',', '.'))


class LoadDictFromFileTests(unittest.TestCase):
    __data = {'2': {'#': '1', 'first': '1\n1', 'second': '22,2', 'third': '', 'int': '123', 'float': '12.3'},
              '3': {'#': '2', 'first': '', 'second': '12345678901234567890', 'third': '"4""4', 'int': '',
                    'float': '12.3'}}
    __data_main = {'1': {'#': '1', 'first': '1\n1', 'second': '22,2', 'third': '', 'int': '123', 'float': '12.3'},
                   '2': {'#': '2', 'first': '', 'second': '12345678901234567890', 'third': '"4""4', 'int': '',
                         'float': '12.3'}}
    __data_recogn = {'2': {'#': 1, 'first': '1\n1', 'second': 22.2, 'third': '', 'int': 123, 'float': 12.3},
                     '3': {'#': 2, 'first': '', 'second': 1.2345678901234567e+19, 'third': '"4""4', 'int': '',
                           'float': 12.3}}
    __data_optim = {'2': {'#': '1', 'first': '1 1', 'second': '22,2', 'third': '', 'int': '123', 'float': '12.3'},
                    '3': {'#': '2', 'first': '', 'second': '12345678901234567890', 'third': '"4""4', 'int': '',
                          'float': '12.3'}}

    def test_csv_import(self):
        test_dict = load('files_for_tests/test_import.csv')
        self.assertEqual(self.__data, test_dict)
        test_dict = load('files_for_tests/test_import.csv', maincolumn='#')
        self.assertEqual(self.__data_main, test_dict)
        test_dict = load('files_for_tests/test_import.csv', recognize=True)
        self.assertEqual(self.__data_recogn, test_dict)
        test_dict = load('files_for_tests/test_import.csv', optimize=True)
        self.assertEqual(self.__data_optim, test_dict)

    def test_xls_import(self):
        test_dict = load('files_for_tests/test_import.xls')
        self.assertEqual(self.__data, test_dict)
        test_dict = load('files_for_tests/test_import.xls', maincolumn='#')
        self.assertEqual(self.__data_main, test_dict)
        test_dict = load('files_for_tests/test_import.xls', recognize=True)
        self.assertEqual(self.__data_recogn, test_dict)
        test_dict = load('files_for_tests/test_import.xls', optimize=True)
        self.assertEqual(self.__data_optim, test_dict)

    def test_xlsx_import(self):
        test_dict = load('files_for_tests/test_import.xlsx')
        self.assertEqual(self.__data, test_dict)
        test_dict = load('files_for_tests/test_import.xlsx', maincolumn='#')
        self.assertEqual(self.__data_main, test_dict)
        test_dict = load('files_for_tests/test_import.xlsx', recognize=True)
        self.assertEqual(self.__data_recogn, test_dict)
        test_dict = load('files_for_tests/test_import.xlsx', optimize=True)
        self.assertEqual(self.__data_optim, test_dict)

    def test_csv_in_and_out(self):
        import SaveDictToFile
        tmp = SaveDictToFile.save_to_csv(self.__data)
        self.assertEqual(self.__data, load(tmp))
        self.assertEqual(self.__data_main, load(tmp, maincolumn='#'))
        self.assertEqual(self.__data_recogn, load(tmp, recognize=True))
        self.assertEqual(self.__data_optim, load(tmp, optimize=True))
        os.remove(tmp)

    def test_xlsx_in_and_out(self):
        import SaveDictToFile
        tmp = SaveDictToFile.save_to_xlsx(self.__data)
        self.assertEqual(self.__data, load(tmp))
        self.assertEqual(self.__data_main, load(tmp, maincolumn='#'))
        self.assertEqual(self.__data_recogn, load(tmp, recognize=True))
        self.assertEqual(self.__data_optim, load(tmp, optimize=True))
        os.remove(tmp)


if __name__ == '__main__':
    unittest.main()
