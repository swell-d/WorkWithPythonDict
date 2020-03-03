import codecs
import copy
import datetime
import os
import pathlib
import re
import unittest

import openpyxl

import GlobalFunctions
from GlobalFunctions import print  # Todo use logging


def __init(data, filename, fieldnames, optimize):
    if fieldnames is None: fieldnames = []
    if not isinstance(data, (dict, list)): raise ValueError('Wrong data')
    if not isinstance(filename, str): raise ValueError('Wrong filename')
    if not isinstance(fieldnames, list): raise ValueError('Wrong fieldnames')
    for key in fieldnames:
        if not isinstance(key, str): raise ValueError('Wrong fieldnames')
    data = copy.deepcopy(data)
    fieldnames = __generate_fieldnames_optimized(data, fieldnames) if optimize else __generate_fieldnames_all(data)
    return data, fieldnames


def __generate_fieldnames_optimized(data, fieldnames):
    new_fieldnames = []
    for each in data.values() if isinstance(data, dict) else data:
        if not isinstance(each, dict): raise ValueError('Wrong data')
        for key, value in each.items():
            if value != '' and key not in new_fieldnames: new_fieldnames.append(str(key))
    additional_fields = [x for x in new_fieldnames if x not in fieldnames]
    cleared_fields = [x for x in fieldnames if x not in new_fieldnames]
    if cleared_fields: print('deleted columns: ' + ', '.join(cleared_fields))
    return [x for x in fieldnames if x in new_fieldnames] + additional_fields


def __generate_fieldnames_all(data):
    all_fieldnames = []
    for each in data.values() if isinstance(data, dict) else data:
        if not isinstance(each, dict): raise ValueError('Wrong data')
        for key, value in each.items():
            if key not in all_fieldnames: all_fieldnames.append(str(key))
    return all_fieldnames


def __view_enhancement(ws):
    def auto_column_width():
        dims = {}
        for row in ws.iter_rows():
            for cell in row:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 1), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 40)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = ws['A2']
    auto_column_width()


@GlobalFunctions.print_run_time
def save_to_xlsx(data, filename='', fieldnames=None, optimize=False, open=False, date_insert=True):
    if not __check_data(data, filename): return None
    data, fieldnames = __init(data, filename, fieldnames, optimize)
    newfilename = _get_new_file_name_with_datetime('.xlsx', filename, date_insert)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    i = -1
    for i, each in enumerate(data.values() if isinstance(data, dict) else data):
        line = []
        for key in fieldnames:
            value = each.get(key, '')
            if not isinstance(value, (int, float)) or len(str(value)) > 10:
                value = str(value) if not optimize else re.sub(r'\s+', ' ', str(value)).strip()
            line.append(value)
        try:
            ws.append(line)
        except openpyxl.utils.exceptions.IllegalCharacterError:
            print(f'save_to_xlsx: IllegalCharacterError: {line}')  # Todo only for debug
            ws.append([x.encode('unicode_escape').decode('utf-8') for x in line])
    __view_enhancement(ws)
    wb.save(newfilename)
    print(f"{newfilename} / {i + 1} lines saved / ", end='')
    if open: os.startfile(newfilename)
    return newfilename


def __check_data(data, filename):
    if data: return True
    print(f'{filename} / nothing to save / ', end='')
    return False


def _get_new_file_name_with_datetime(filetype, filename, date_insert):
    if not date_insert:
        if '\\' in filename: pathlib.Path(filename[:filename.rfind('\\')]).mkdir(parents=True, exist_ok=True)
        return f'{filename}{filetype}'
    date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d_%H-%M")
    if '\\' not in filename: return f'{date}_{filename}{filetype}'
    last_pos = filename.rfind('\\')
    pathlib.Path(filename[:last_pos]).mkdir(parents=True, exist_ok=True)
    return f'{filename[:last_pos]}\\{date}_{filename[last_pos + 1:]}{filetype}'


@GlobalFunctions.print_run_time
def save_to_csv(data, filename='', fieldnames=None, optimize=False, open=False, date_insert=True):
    if not __check_data(data, filename): return None
    SEP, QC, NL = ',', '"', '\r\n'  # separator, quote char, new line
    data, fieldnames = __init(data, filename, fieldnames, optimize)
    newfilename = _get_new_file_name_with_datetime('.csv', filename, date_insert)
    with codecs.open(newfilename, 'w', encoding='utf-8') as file:
        file.write(SEP.join([f'{QC}{x}{QC}' for x in fieldnames]) + NL)
        i = -1
        for i, each in enumerate(data.values() if isinstance(data, dict) else data):
            line = []
            for key in fieldnames:
                value = each.get(key, '')
                if isinstance(value, float): value = str(value)
                value = str(value) if not optimize else re.sub(r'\s+', ' ', str(value)).strip()
                line.append(value.replace('"', '""'))
            file.write(SEP.join([f'{QC}{x}{QC}' for x in line]) + NL)
    print(f"{newfilename} / {i + 1} lines saved / ", end='')
    if open: os.startfile(newfilename)
    return newfilename


def save_to_files(data, filename='', fieldnames=None, optimize=False, open=False, date_insert=True):
    return (save_to_xlsx(data, filename, fieldnames, optimize, open, date_insert),
            save_to_csv(data, filename, fieldnames, optimize, open, date_insert))


def _param_list_extend():
    return [
        # main
        'url',
        'sku',
        'manufacturer',  # may be ident
        'herstellernummer',
        'category_ids',
        'name',
        'short_description',
        'description',

        # images
        'image',
        'small_image',
        'thumbnail',
        'media_gallery',
        'logo',

        # dop
        'weight',
        'ean',

        # same_for_all
        'url_key',
        'status',  # ident
        'qty',  # ident
        'is_in_stock',  # ident
        'tax_class_id',  # ident
        'attribute_set_name',  # ident

        # specific
        'websites',  # ident
        'lieferant',  # ident
        'am_shipping_type',  # ident
        'country_of_manufacture',  # may be ident
        'delivery_time',
        'relation',  # через ||

        # price
        'price',
        'special_price'
        # 'original_price',
    ]


class SaveDictToFileTests(unittest.TestCase):
    __data = {'2': {'first': '1\r\n1', 'second': '22.2', 'third': ''},
              '3': {'first': '', 'second': '12345678901234567890', 'third': '"4""4'}}

    def test_save_to_xlsx(self):
        result = [['first', 'second', 'third'], ['1\r\n1', '22.2', 'None'],
                  ['None', '12345678901234567890', '"4""4']]
        xlsx = []
        file_name = save_to_xlsx(self.__data)
        sheet = openpyxl.load_workbook(file_name).active
        for row in range(1, sheet.max_row + 1):
            row_data = []
            for column in range(1, sheet.max_column + 1):
                row_data.append(str(sheet.cell(row=row, column=column).value))
            xlsx.append(row_data)
        self.assertEqual(result, xlsx)
        import LoadDictFromFile
        self.assertEqual(self.__data, LoadDictFromFile.load(file_name))
        os.remove(file_name)

    def test_save_to_csv(self):
        result = '"first","second","third"\r\n"1\r\n1","22.2",""\r\n"","12345678901234567890","""4""""4"\r\n'
        file_name = save_to_csv(self.__data)
        with codecs.open(file_name, 'r', encoding='utf-8') as file:
            csv = file.read()
        self.assertEqual(result, csv)
        import LoadDictFromFile
        self.assertEqual(self.__data, LoadDictFromFile.load(file_name))
        os.remove(file_name)

    def test_get_new_file_name(self):
        date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d_%H-%M")
        self.assertEqual('filename.xlsx',
                         _get_new_file_name_with_datetime('.xlsx', 'filename', date_insert=False))
        self.assertEqual(f'{date}_filename.xlsx',
                         _get_new_file_name_with_datetime('.xlsx', 'filename', date_insert=True))
        self.assertEqual('C:\\tmp\\filename.xlsx',
                         _get_new_file_name_with_datetime('.xlsx', 'C:\\tmp\\filename', date_insert=False))
        self.assertEqual(f'C:\\tmp\\{date}_filename.xlsx',
                         _get_new_file_name_with_datetime('.xlsx', 'C:\\tmp\\filename', date_insert=True))


if __name__ == '__main__':
    unittest.main()
